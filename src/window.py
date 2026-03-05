"""
GUI-модуль для генерации, редактирования и визуализации учебного расписания.

Модуль содержит следующие основные компоненты:

ScheduleGeneratorWorkerThread
    Фоновый поток, выполняющий многократную генерацию расписаний,
    выбирающий лучшее по рейтингу и передающий результат в основной поток.

ScheduleGeneratorDialog
    Диалоговое окно, позволяющее выбрать количество итераций и следить
    за прогрессом генерации расписания.

InputDataDialog
    Диалог для просмотра и редактирования всех исходных данных:
    групп, дисциплин, преподавателей, аудиторий и их доступности.

ErrorDialog
    Окно отображения ошибок генерации — тех дисциплин и групп,
    для которых не удалось поставить пары, а также оставшихся часов.

MainWindow
    Главное окно приложения. Позволяет запускать генерацию,
    просматривать расписание, выгружать его в Excel, редактировать входные данные
    и просматривать ошибки.

Модуль обеспечивает полный графический интерфейс для работы с расписанием:
от ввода данных и фоновой оптимизационной генерации
до отображения результата и обработки ошибок.
"""

import copy
import datetime
import os
import sys
import time
import traceback
import typing
from pathlib import Path
from pprint import pp

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PyQt5.QtCore import QEvent, Qt, QThread, QTimer, pyqtSignal
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QComboBox,
    QDialog,
    QHBoxLayout,
    QLabel,
    QListWidget,
    QMainWindow,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSlider,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from config.constants import (
    DAY_MAPPING,
    DEFAULT_ITERATIONS,
    EXCEL_COLUMN_WIDTH_PADDING,
    HELP_LABEL_WIDTH,
    MAX_ITERATIONS,
    MIN_ITERATIONS,
    SECONDS_PER_MINUTE,
    SHIFT_ONE,
    SHIFT_THREE,
    SHIFT_TWO,
    WINDOW_HEIGHT,
    WINDOW_WIDTH,
    WINDOW_X,
    WINDOW_Y,
    ExportError,
)
from config.logger import get_logger
from config.messages import (
    QMessageBoxDataErrors,
    QMessageBoxHelpTexts,
    QMessageBoxMessages,
    QMessageBoxTitles,
)
from src import best_of, db, schedule_maker

logger = get_logger("window")


class ScheduleGeneratorWorkerThread(QThread):
    result_ready = pyqtSignal(schedule_maker.Schedule, dict)

    def __init__(self, iterations: int, data):  # noqa: ANN001
        super().__init__()
        self.iterations = iterations
        self.data = data
        self.best_data = None
        self.best_schedule_obj = None
        self.best_rating = float('-inf')
        self.best_schedule_counts = None
        self.progress_value = 0
        self.remaining_time = 0
        self.is_running = True

    def run(self):
        logger.info(
            "Starting schedule generation thread (%d iterations)", self.iterations
        )
        start_time = time.time()

        for iteration in range(1, self.iterations + 1):
            if not self.is_running:
                logger.info("Generation thread stopped")
                break
            logger.debug("Iteration %d/%d", iteration, self.iterations)
            # Обновляем только переменные прогресса
            passed_time = time.time() - start_time
            approx_time = self.iterations * passed_time / iteration
            self.remaining_time = approx_time - passed_time
            self.progress_value = round((iteration / self.iterations) * 100, 2)

            # Основная работа
            data_copy = best_of.shuffle_data(self.data)
            current_schedule_obj = schedule_maker.make_full_schedule(data_copy)
            schedule_rating = best_of.rate_schedule(
                current_schedule_obj.pairs,
                data_copy,
                current_schedule_obj.remaining_data,
            )

            if schedule_rating > self.best_rating:
                self.best_rating = schedule_rating
                del self.best_schedule_obj
                self.best_schedule_obj = copy.deepcopy(current_schedule_obj)
                self.best_data = copy.deepcopy(data_copy)
            else:
                del current_schedule_obj

        if not self.is_running:
            return

        # Завершение работы и отправка результата
        self.best_schedule_counts = best_of.get_counts(
            self.best_schedule_obj.pairs,
            self.best_data,
            self.best_schedule_obj.remaining_data,
        )
        rating = {
            "rate": self.best_rating,
            "teachers_gaps_count": self.best_schedule_counts["teachers_gaps_count"],
            "offline_pairs_gaps": self.best_schedule_counts["offline_pairs_gaps"],
            "overworked_teachers": self.best_schedule_counts["overworked_teachers"],
            "unissued_hours": self.best_schedule_counts["unissued_hours"],
        }
        logger.info("Generation thread finished, best rating: %.2f", self.best_rating)
        self.result_ready.emit(self.best_schedule_obj, rating)
        self.stop()

    def stop(self):
        logger.debug("Stopping generation thread")
        self.is_running = False
        self.quit()


class ScheduleGeneratorDialog(QDialog):
    result_obtained = pyqtSignal(schedule_maker.Schedule, dict)

    def __init__(self, data):  # noqa: ANN001
        logger.debug("Creating schedule generation dialog")
        super().__init__()
        self.data = data
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self.update_progress_bar)
        self.worker_thread = None

        self.result = None
        self.rating = None

        with Path(
            db.resource_path(f"../css/{self.__class__.__name__}.css")
        ).open() as style_file:
            self.setStyleSheet(style_file.read())
        self.setWindowIcon(QIcon(db.resource_path("../icon.ico")))

        self.setWindowTitle("Генератор лучшего расписания")
        self.setGeometry(WINDOW_X, WINDOW_Y, 400, 300)

        layout = QVBoxLayout()

        self.worker_thread = None

        self.number_label = QLabel(f"Выберите число итераций: {DEFAULT_ITERATIONS}")
        layout.addWidget(self.number_label)

        self.number_slider = QSlider(Qt.Horizontal)
        self.number_slider.setMinimum(MIN_ITERATIONS)
        self.number_slider.setMaximum(MAX_ITERATIONS)
        self.number_slider.setValue(DEFAULT_ITERATIONS)
        self.number_slider.update()
        self.number_slider.setTickInterval(1000)
        self.number_slider.valueChanged.connect(self.update_number_label)
        layout.addWidget(self.number_slider)

        self.generate_button = QPushButton("Генерация")
        self.generate_button.clicked.connect(self.start_generation)
        layout.addWidget(self.generate_button)

        self.progress_bar = QProgressBar(self)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)

        self.remaining_time_label = QLabel()
        layout.addWidget(self.remaining_time_label)

        self.setLayout(layout)

    @typing.override
    def closeEvent(self, event):  # noqa: ANN001
        if self.worker_thread and self.worker_thread.isRunning():
            self.worker_thread.stop()
        super().closeEvent(event)

    def start_generation(self):
        self.generate_button.setEnabled(False)
        self.progress_bar.setValue(0)
        selected_number = self.number_slider.value()

        # Запуск рабочего потока и таймера
        self.worker_thread = ScheduleGeneratorWorkerThread(selected_number, self.data)
        self.worker_thread.result_ready.connect(self.handle_result)
        self.worker_thread.finished.connect(self.on_task_finished)
        self.worker_thread.start()

        self.update_timer.start(1000)

    def update_progress_bar(self):
        # Обновление состояния интерфейса на основе данных из потока
        if self.worker_thread:
            self.progress_bar.setValue(int(self.worker_thread.progress_value))
            remaining_time = self.worker_thread.remaining_time
            self.remaining_time_label.setText(
                f"Осталось времени: {str(round(remaining_time // 59)) + 'м, ' if remaining_time >= SECONDS_PER_MINUTE else ''}"
                f"{round(remaining_time % 59)!s}с."
            )

    def update_number_label(self):
        self.number_label.setText(
            f"Выберите число итераций:\n{self.number_slider.value()}"
        )

    def handle_result(self, result, rating):  # noqa: ANN001
        self.result = result
        self.rating = rating
        self.result_obtained.emit(result, rating)

    def on_task_finished(self):
        self.generate_button.setEnabled(True)
        self.accept()

    def get_result(self):
        return self.result, self.rating

    def event(self, event):  # noqa: ANN001
        if event.type() == QEvent.Type(124):
            self.show_help()
            return True
        return super().event(event)

    def show_help(self):
        # Отображение справки
        QMessageBox.information(
            self,
            QMessageBoxTitles.HELP.value,
            QMessageBoxMessages.SCHEDULE_GENERATOR_HELP.value,
        )


class InputDataDialog(QDialog):
    def __init__(self, data=None):  # noqa: ANN001
        logger.debug("Creating input data dialog")
        super().__init__(parent=None)

        with Path(
            db.resource_path(f"../css/{self.__class__.__name__}.css")
        ).open() as f:
            self.setStyleSheet(f.read())
        self.setWindowIcon(QIcon(db.resource_path("../icon.ico")))

        self.vars_to_redact = {
            "Смены групп": "groups_shift",
            "КУГ": "discipline_hours",
            "Преподаватели": "teachers",
            "Аудитории": "rooms",
            "Расписание\nпреподавателей": "teachers_work_hours",
            "Расписание\nаудиторий": "rooms_availability_hours",
        }

        self.setWindowTitle("Ввод данных")
        self.setGeometry(WINDOW_X, WINDOW_Y, WINDOW_WIDTH, WINDOW_HEIGHT)
        self.setWindowFlags(Qt.Window)

        self.layout = QHBoxLayout(self)

        self.variable_list = QListWidget()
        self.variable_list.setFixedWidth(250)
        self.variable_list.addItems(list(self.vars_to_redact.keys()))
        self.variable_list.currentItemChanged.connect(self.display_variable_data)
        self.layout.addWidget(self.variable_list)

        self.data_table = QTableWidget()
        self.layout.addWidget(self.data_table)

        self.button_layout = QVBoxLayout()

        self.add_row_button = QPushButton("Добавить строку")
        self.add_row_button.clicked.connect(self.add_row)
        self.button_layout.addWidget(self.add_row_button)

        self.delete_selected_rows_button = QPushButton("Удалить выделенные строки")
        self.delete_selected_rows_button.clicked.connect(self.delete_selected_rows)
        self.button_layout.addWidget(self.delete_selected_rows_button)

        self.save_button = QPushButton("Сохранить изменения")
        self.save_button.clicked.connect(self.save_changes)
        self.button_layout.addWidget(self.save_button)

        self.clear_data_button = QPushButton("Очистить все данные")
        self.clear_data_button.clicked.connect(self.clear_data)
        self.button_layout.addWidget(self.clear_data_button)

        self.load_test_data_button = QPushButton("Загрузить тестовые данные")
        self.load_test_data_button.clicked.connect(self.load_test_data)
        self.button_layout.addWidget(self.load_test_data_button)

        self.back_button = QPushButton("Назад")
        self.back_button.clicked.connect(self.close)
        self.button_layout.addWidget(self.back_button)

        self.var_help_label = QLabel()
        self.var_help_label.setFixedWidth(HELP_LABEL_WIDTH)
        self.var_help_label.setWordWrap(True)
        self.button_layout.addWidget(self.var_help_label)

        self.layout.addLayout(self.button_layout)

        self.current_variable = None

        # Use passed data or load from database
        if data is not None:
            self.data = data
        # Initialize data attribute - load from database or create empty data
        elif db.check_exists_data():
            self.data = db.load_data()
        else:
            self.data = db.EmptyData()

        # Select the first item in the variable list
        if self.variable_list.count() > 0:
            self.variable_list.setCurrentRow(0)

    @typing.override
    def closeEvent(self, event):  # noqa: ANN001
        # Auto-save data when dialog is closed
        try:
            db.save_data(self.data)
            logger.info("Data automatically saved on dialog close")
        except Exception:
            logger.exception("Error during auto save")
        event.accept()

    def load_test_data(self):
        resp = QMessageBox.question(
            self,
            QMessageBoxTitles.CONFIRMATION.value,
            QMessageBoxMessages.LOAD_TEST_DATA_CONFIRMATION.value,
            QMessageBox.Yes | QMessageBox.No,
        )
        if resp == QMessageBox.Yes:
            self.data = db.ExampleData()
            self.display_variable_data(self.data_table.currentItem())

    def clear_data(self):
        resp = QMessageBox.question(
            self,
            QMessageBoxTitles.CONFIRMATION.value,
            QMessageBoxMessages.CLEAR_DATA_CONFIRMATION.value,
            QMessageBox.Yes | QMessageBox.No,
        )
        if resp == QMessageBox.Yes:
            self.data = db.EmptyData()
            self.display_variable_data(self.data_table.currentItem())

    def delete_selected_rows(self):
        resp = QMessageBox.question(
            self,
            QMessageBoxTitles.CONFIRMATION.value,
            QMessageBoxMessages.DELETE_ROWS_CONFIRMATION.value,
            QMessageBox.Yes | QMessageBox.No,
        )
        if resp != QMessageBox.Yes:
            return

        selected_indexes = self.data_table.selectionModel().selectedRows()
        selected_rows = sorted(
            [index.row() for index in selected_indexes], reverse=True
        )
        for row in selected_rows:
            self.data_table.model().removeRow(row)

    def add_row(self):
        row = self.data_table.rowCount()
        self.data_table.insertRow(row)

        if self.current_variable == "discipline_hours":
            dropdown = QComboBox()
            dropdown.addItems(self.data.groups_shift.keys())
            self.data_table.setCellWidget(row, 0, dropdown)
        elif self.current_variable == "groups_shift":
            self.data_table.setItem(row, 0, QTableWidgetItem("Группа"))
            self.data_table.setItem(row, 1, QTableWidgetItem("Номер смены"))
        elif self.current_variable == "teachers":
            self.data_table.setItem(row, 0, QTableWidgetItem("ФИО"))
            self.data_table.setItem(
                row, 1, QTableWidgetItem("Дисциплины через запятую")
            )
            self.data_table.setItem(row, 2, QTableWidgetItem("Группы через запятую"))
        elif self.current_variable == "rooms":
            self.data_table.setItem(row, 0, QTableWidgetItem("Аудитория"))
            self.data_table.setItem(row, 1, QTableWidgetItem("Да / Нет"))
        elif self.current_variable == "teachers_work_hours":
            self.data_table.setItem(row, 0, QTableWidgetItem("ФИО"))
        elif self.current_variable == "rooms_availability_hours":
            self.data_table.setItem(row, 0, QTableWidgetItem("Аудитория"))

        if self.current_variable not in [
            "teachers_work_hours",
            "rooms_availability_hours",
        ]:
            self.data_table.resizeColumnsToContents()
            return
        num_pairs = len(self.data.teachers_schedule_time)
        days_of_week = list(self.data.days)

        for col, _day in enumerate(days_of_week, start=1):
            day_schedule = [False] * num_pairs

            cell_widget = QWidget()
            cell_layout = QHBoxLayout(cell_widget)
            cell_layout.setContentsMargins(0, 0, 0, 0)

            for slot in day_schedule:
                check_box = QCheckBox()
                check_box.setChecked(slot)
                cell_layout.addWidget(check_box)

            cell_widget.setLayout(cell_layout)
            self.data_table.setCellWidget(row, col, cell_widget)

        self.data_table.resizeColumnsToContents()

    def display_variable_data(self, current):  # noqa: ANN001
        if not current:
            return

        variable_name = self.vars_to_redact[current.text()]
        self.current_variable = variable_name
        variable_data = getattr(self.data, variable_name)

        self.data_table.setRowCount(0)

        if variable_name == "groups_shift":
            self._display_groups_shift(variable_data)
        elif variable_name == "discipline_hours":
            self._display_discipline_hours(variable_data)
        elif variable_name == "teachers":
            self._display_teachers(variable_data)
        elif variable_name == "rooms":
            self._display_rooms(variable_data)
        elif variable_name in ["teachers_work_hours", "rooms_availability_hours"]:
            self._display_schedule_data(variable_name, variable_data)

        self.data_table.setEditTriggers(QTableWidget.DoubleClicked)
        self.data_table.resizeColumnsToContents()

    def _display_groups_shift(self, variable_data: dict) -> None:
        """Display groups shift data in table."""
        self.var_help_label.setText(QMessageBoxHelpTexts.GROUPS_SHIFT_HELP.value)
        headers = ["Группа", "Смена"]
        self.data_table.setColumnCount(len(headers))
        self.data_table.setHorizontalHeaderLabels(headers)

        for row, (group, schedule) in enumerate(variable_data.items()):
            shift = self._determine_shift(schedule)
            self.data_table.insertRow(row)
            self.data_table.setItem(row, 0, QTableWidgetItem(group))
            self.data_table.setItem(row, 1, QTableWidgetItem(shift))

    def _determine_shift(self, schedule: dict) -> str:
        """Determine shift number based on schedule."""
        if not schedule or 1 not in schedule:
            return None

        first_pair = schedule[1]

        # Check all shifts independently
        if 1 in self.data.schedule_time_shift_1:
            shift1_first = self.data.schedule_time_shift_1[1]
            if (
                first_pair.start == shift1_first.start
                and first_pair.end == shift1_first.end
            ):
                return "1"

        if 1 in self.data.schedule_time_shift_2:
            shift2_first = self.data.schedule_time_shift_2[1]
            if (
                first_pair.start == shift2_first.start
                and first_pair.end == shift2_first.end
            ):
                return "2"

        if 1 in self.data.schedule_time_shift_3:
            shift3_first = self.data.schedule_time_shift_3[1]
            if (
                first_pair.start == shift3_first.start
                and first_pair.end == shift3_first.end
            ):
                return "3"

        return None

    def _display_discipline_hours(self, variable_data: dict) -> None:
        """Display discipline hours data in table."""
        self.var_help_label.setText(QMessageBoxHelpTexts.DISCIPLINE_HOURS_HELP.value)
        self.data_table.setColumnCount(3)
        self.data_table.setHorizontalHeaderLabels(["Группа", "Дисциплина", "Часы"])

        row = 0
        for group, disciplines in variable_data.items():
            for discipline, hours in disciplines.items():
                self.data_table.insertRow(row)
                self.data_table.setItem(row, 0, QTableWidgetItem(group))
                self.data_table.setItem(row, 1, QTableWidgetItem(discipline))
                self.data_table.setItem(row, 2, QTableWidgetItem(str(hours)))
                row += 1

    def _display_teachers(self, variable_data: dict) -> None:
        """Display teachers data in table."""
        self.var_help_label.setText(QMessageBoxHelpTexts.TEACHERS_HELP.value)
        self.data_table.setColumnCount(3)
        self.data_table.setHorizontalHeaderLabels(["Имя", "Дисциплины", "Группы"])

        row = 0
        for name, presets in variable_data.items():
            for teacher in presets:
                self.data_table.insertRow(row)
                self.data_table.setItem(row, 0, QTableWidgetItem(name))
                self.data_table.setItem(
                    row, 1, QTableWidgetItem(", ".join(teacher.disciplines))
                )
                self.data_table.setItem(
                    row, 2, QTableWidgetItem(", ".join(teacher.groups))
                )
                row += 1

    def _display_rooms(self, variable_data: dict) -> None:
        """Display rooms data in table."""
        self.var_help_label.setText(QMessageBoxHelpTexts.ROOMS_HELP.value)
        self.data_table.setColumnCount(2)
        self.data_table.setHorizontalHeaderLabels(["Аудитория", "Онлайн"])

        for row, (room, room_obj) in enumerate(variable_data.items()):
            self.data_table.insertRow(row)
            self.data_table.setItem(row, 0, QTableWidgetItem(room))
            self.data_table.setItem(
                row, 1, QTableWidgetItem("Да" if room_obj.is_online else "Нет")
            )

    def _display_schedule_data(self, variable_name: str, variable_data: dict) -> None:
        """Display schedule data for teachers or rooms."""
        if variable_name == "teachers_work_hours":
            self.var_help_label.setText(
                QMessageBoxHelpTexts.TEACHERS_WORK_HOURS_HELP.value
            )
        else:
            self.var_help_label.setText(
                QMessageBoxHelpTexts.ROOMS_AVAILABILITY_HELP.value
            )
        self._setup_schedule_table(variable_data)

    def _setup_schedule_table(self, schedule_data) -> None:  # noqa: ANN001
        days_of_week = list(self.data.days)
        num_pairs = len(self.data.teachers_schedule_time)

        # Устанавливаем столбцы: один для имени, остальные для расписания по дням
        self.data_table.setColumnCount(1 + len(days_of_week))
        headers = ["Имя", *days_of_week]
        self.data_table.setHorizontalHeaderLabels(headers)

        for row, (name, schedule) in enumerate(schedule_data.items()):
            self.data_table.insertRow(row)
            self.data_table.setItem(row, 0, QTableWidgetItem(name))

            for col, day in enumerate(days_of_week, start=1):
                day_schedule = schedule.schedule_for_days.get(day, [False] * num_pairs)

                cell_widget = QWidget()
                cell_layout = QHBoxLayout(cell_widget)
                cell_layout.setContentsMargins(0, 0, 0, 0)

                # Создаем по одному QCheckBox для каждой пары в день
                for slot in day_schedule:
                    check_box = QCheckBox()
                    check_box.setChecked(slot)
                    cell_layout.addWidget(check_box)

                cell_widget.setLayout(cell_layout)
                self.data_table.setCellWidget(row, col, cell_widget)

    def _save_schedule_changes(self, schedule_data) -> None:  # noqa: ANN001
        for row in range(self.data_table.rowCount()):
            name = self.data_table.item(row, 0).text()

            updated_schedule = {}
            for col in range(1, self.data_table.columnCount()):
                day = self.data_table.horizontalHeaderItem(col).text()

                cell_widget = self.data_table.cellWidget(row, col)
                day_schedule = [
                    check_box.isChecked()
                    for check_box in cell_widget.findChildren(QCheckBox)
                ]

                updated_schedule[day] = day_schedule

            if name in schedule_data:
                schedule_data[name].schedule_for_days = updated_schedule
            elif self.current_variable == "teachers_work_hours":
                # Создаем TeachersSchedule с правильным порядком параметров
                params = []
                num_pairs = len(self.data.teachers_schedule_time)
                for eng_field in ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]:
                    rus_name = DAY_MAPPING[eng_field]
                    day_data = updated_schedule.get(rus_name)
                    if day_data is None:
                        day_data = [False] * num_pairs
                    params.append(day_data)
                schedule_data[name] = db.TeachersSchedule(*params)
            elif self.current_variable == "rooms_availability_hours":
                # Создаем RoomSchedule с правильным порядком параметров
                params = []
                num_pairs = len(self.data.teachers_schedule_time)
                for eng_field in ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]:
                    rus_name = DAY_MAPPING[eng_field]
                    day_data = updated_schedule.get(rus_name)
                    if day_data is None:
                        day_data = [False] * num_pairs
                    params.append(day_data)
                schedule_data[name] = db.RoomSchedule(*params)

    def save_changes(self):
        """Save changes made to the current variable data."""
        if self.current_variable is None:
            return

        invalid_data = []
        table_name = self.current_variable
        ru_table_name = self._get_table_display_name(table_name)

        try:
            variable_data = getattr(self.data, table_name)

            if table_name == "groups_shift":
                variable_data = self._save_groups_shift(variable_data, invalid_data)
            elif table_name == "discipline_hours":
                variable_data = self._save_discipline_hours(variable_data, invalid_data)
            elif table_name == "teachers":
                variable_data = self._save_teachers(variable_data, invalid_data)
            elif table_name == "rooms":
                variable_data = self._save_rooms(variable_data, invalid_data)
            elif table_name in ["teachers_work_hours", "rooms_availability_hours"]:
                self._save_schedule_data(variable_data)

            setattr(self.data, table_name, variable_data)
            self._handle_save_result(invalid_data, ru_table_name)

        except AttributeError:
            self._show_save_empty_fields_error()
        except (OSError, ValueError, TypeError) as e:
            self._show_critical_save_error(e)

    def _get_table_display_name(self, table_name: str) -> str:
        """Get the display name for the table."""
        return next(
            (key for key, value in self.vars_to_redact.items() if value == table_name),
            None,
        )

    def _save_groups_shift(self, variable_data: dict, invalid_data: list) -> dict:
        """Save groups shift data."""
        variable_data = {}  # Reset variable

        for row in range(self.data_table.rowCount()):
            try:
                group = self.data_table.item(row, 0).text()
                shift = int(self.data_table.item(row, 1).text())

                schedule = self._get_schedule_by_shift(shift)
                variable_data[group] = schedule

            except (ValueError, AttributeError, IndexError):
                invalid_data.append(self._get_row_data(row))

        return variable_data

    def _get_schedule_by_shift(self, shift: int) -> dict:
        """Get schedule data by shift number."""
        if shift == SHIFT_ONE:
            return self.data.schedule_time_shift_1
        if shift == SHIFT_TWO:
            return self.data.schedule_time_shift_2
        if shift == SHIFT_THREE:
            return self.data.schedule_time_shift_3
        msg = "Неверный номер смены"
        raise ValueError(msg)

    def _save_discipline_hours(self, variable_data: dict, invalid_data: list) -> dict:
        """Save discipline hours data."""
        variable_data = {}

        for row in range(self.data_table.rowCount()):
            try:
                group = self._get_group_from_cell(row)
                discipline = self.data_table.item(row, 1).text()
                hours = int(self.data_table.item(row, 2).text())

                if group not in variable_data:
                    variable_data[group] = {}
                variable_data[group][discipline] = hours

            except (ValueError, AttributeError):
                invalid_data.append(self._get_row_data(row))

        return variable_data

    def _get_group_from_cell(self, row: int) -> str:
        """Get group name from table cell."""
        if self.data_table.item(row, 0) is None:
            return self.data_table.cellWidget(row, 0).currentText()
        return self.data_table.item(row, 0).text()

    def _save_teachers(self, variable_data: dict, invalid_data: list) -> dict:
        """Save teachers data."""
        variable_data = {}

        for row in range(self.data_table.rowCount()):
            try:
                name = self.data_table.item(row, 0).text()
                disciplines = set(self.data_table.item(row, 1).text().split(", "))
                groups = set(self.data_table.item(row, 2).text().split(", "))

                if name not in variable_data:
                    variable_data[name] = [db.Teacher(name, disciplines, groups)]
                else:
                    variable_data[name].append(db.Teacher(name, disciplines, groups))

                if name not in self.data.teachers_work_hours:
                    self.data.teachers_work_hours[name] = db.TeachersSchedule()

            except AttributeError:
                invalid_data.append(self._get_row_data(row))

        self._cleanup_removed_teachers(variable_data)
        pp(variable_data)
        return variable_data

    def _cleanup_removed_teachers(self, variable_data: dict) -> None:
        """Remove teachers that are no longer in the data."""
        data_copy = copy.deepcopy(self.data.teachers_work_hours)
        for teacher in self.data.teachers_work_hours:
            if teacher not in variable_data:
                data_copy.pop(teacher)
        self.data.teachers_work_hours = data_copy

    def _save_rooms(self, variable_data: dict, invalid_data: list) -> dict:
        """Save rooms data."""
        variable_data = {}

        for row in range(self.data_table.rowCount()):
            try:
                room = self.data_table.item(row, 0).text()
                is_online = self.data_table.item(row, 1).text() == "Да"
                variable_data[room] = db.Room(is_online)

                if room not in self.data.rooms_availability_hours:
                    self.data.rooms_availability_hours[room] = db.RoomSchedule()

            except AttributeError:
                invalid_data.append(self._get_row_data(row))

        self._cleanup_removed_rooms(variable_data)
        return variable_data

    def _cleanup_removed_rooms(self, variable_data: dict) -> None:
        """Remove rooms that are no longer in the data."""
        data_copy = copy.deepcopy(self.data.rooms_availability_hours)
        for room in self.data.rooms_availability_hours:
            if room not in variable_data:
                data_copy.pop(room)
        self.data.rooms_availability_hours = data_copy

    def _save_schedule_data(self, variable_data: dict) -> None:
        """Save schedule data for teachers or rooms."""
        try:
            self._save_schedule_changes(variable_data)
        except OSError as e:  # More specific exception
            QMessageBox.warning(
                self,
                QMessageBoxTitles.ERROR.value,
                QMessageBoxMessages.SCHEDULE_SAVE_ERROR_TEMPLATE.value.format(e),
            )

    def _get_row_data(self, row: int) -> list[str]:
        """Get data from a table row."""
        return [
            self.data_table.item(row, col).text()
            if self.data_table.item(row, col)
            else ""
            for col in range(self.data_table.columnCount())
        ]

    def _handle_save_result(self, invalid_data: list, ru_table_name: str) -> None:
        """Handle the result of save operation."""
        if invalid_data:
            self._show_data_errors(invalid_data, ru_table_name)
        else:
            db.save_data(self.data)
            self._show_save_success(ru_table_name)

    def _show_data_errors(self, invalid_data: list, ru_table_name: str) -> None:
        """Show data validation errors."""
        error_rows = ""
        for i, row_data in enumerate(invalid_data[:3]):
            error_rows += QMessageBoxDataErrors.ROW_ERROR_TEMPLATE.value.format(
                i + 1, ", ".join(row_data)
            )

        error_message = QMessageBoxMessages.DATA_ERROR_TEMPLATE.value.format(
            ru_table_name, error_rows
        )
        QMessageBox.warning(self, QMessageBoxTitles.DATA_ERROR.value, error_message)

    def _show_save_success(self, ru_table_name: str) -> None:
        """Show save success message."""
        QMessageBox.information(
            self,
            QMessageBoxTitles.SAVING.value,
            QMessageBoxMessages.DATA_SAVED_SUCCESS_TEMPLATE.value.format(ru_table_name),
        )

    def _show_save_empty_fields_error(self) -> None:
        """Show error for empty fields."""
        QMessageBox.critical(
            self,
            QMessageBoxTitles.ERROR.value,
            QMessageBoxMessages.SAVE_EMPTY_FIELDS_ERROR.value,
        )

    def _show_critical_save_error(self, e: Exception) -> None:
        """Show critical save error."""
        QMessageBox.critical(
            self,
            QMessageBoxTitles.CRITICAL_ERROR.value,
            QMessageBoxMessages.CRITICAL_SAVE_ERROR_TEMPLATE.value.format(
                type(e).__name__, e, traceback.format_exc()
            ),
        )

    def event(self, event):  # noqa: ANN001
        if event.type() == QEvent.Type(124):
            self.show_help()
            return True
        return super().event(event)

    def show_help(self):
        # Отображение справки
        QMessageBox.information(
            self,
            QMessageBoxTitles.HELP.value,
            QMessageBoxMessages.INPUT_DATA_HELP.value,
        )


class ErrorDialog(QDialog):
    def __init__(self, errors: list, remaining_data: db.Data):
        logger.debug("Creating error dialog, errors count: %d", len(errors))
        super().__init__()
        self.setWindowTitle(QMessageBoxTitles.SCHEDULE_GENERATION_ERRORS.value)
        self.setGeometry(
            WINDOW_X, WINDOW_Y, 800, 500
        )  # Увеличиваем ширину окна для списка оставшихся часов
        self.setWindowFlags(
            self.windowFlags() | Qt.WindowContextHelpButtonHint
        )  # Добавляем кнопку справки

        # Загрузка стиля
        with Path(
            db.resource_path(f"../css/{self.__class__.__name__}.css")
        ).open() as f:
            self.setStyleSheet(f.read())
        self.setWindowIcon(QIcon(db.resource_path("../icon.ico")))

        # Основной макет с горизонтальным расположением для левой и правой частей
        main_layout = QHBoxLayout(self)

        # Левая часть (список ошибок и информация о текущей ошибке)
        left_layout = QVBoxLayout()
        self.errors = errors

        # Список ошибок
        self.error_list_widget = QListWidget()
        self.error_list_widget.addItems(
            [
                f"{i + 1} / Группа: {error.group}, Дисциплина: {error.discipline}"
                for i, error in enumerate(errors)
            ]
        )
        self.error_list_widget.currentItemChanged.connect(self.display_error_info)

        # Заголовок и описание ошибок
        self.help_text = QLabel(
            "Невозможно поставить пары для данных дисциплин и групп:"
        )
        left_layout.addWidget(self.help_text)
        left_layout.addWidget(self.error_list_widget)

        self.group_label = QLabel("| Группа: ")
        self.discipline_label = QLabel("| Дисциплина: ")
        self.hours_label = QLabel("| Оставшиеся часы: ")
        left_layout.addWidget(self.group_label)
        left_layout.addWidget(self.discipline_label)
        left_layout.addWidget(self.hours_label)

        main_layout.addLayout(left_layout)

        # Правая часть (список дисциплин с оставшимися часами)
        right_layout = QVBoxLayout()
        self.remaining_hours_list_widget = QListWidget()

        # Заголовок для оставшихся часов
        remaining_hours_label = QLabel("Дисциплины с оставшимися часами:")
        right_layout.addWidget(remaining_hours_label)
        right_layout.addWidget(self.remaining_hours_list_widget)

        # Добавление данных об оставшихся часах в виджет
        for group, disciplines in remaining_data.discipline_hours.items():
            for discipline, hours in disciplines.items():
                if hours > 0:
                    self.remaining_hours_list_widget.addItem(
                        f"Группа: {group},\nДисциплина: {discipline},\nОсталось часов: {hours}"
                    )

        self.back_button = QPushButton("Назад")
        self.back_button.clicked.connect(self.close)
        right_layout.addWidget(self.back_button)

        main_layout.addLayout(right_layout)

    def display_error_info(self, current):  # noqa: ANN001
        def pair_text(count: int) -> str:
            if count % 10 == 1:
                return "пара"
            if count % 10 in [2, 3, 4]:
                return "пары"
            return "пар"

        if current:
            current_error = self.errors[int(current.text().split("/")[0].strip()) - 1]
            self.group_label.setText(f"| Группа: {current_error.group}")
            self.discipline_label.setText(f"| Дисциплина: {current_error.discipline}")
            self.hours_label.setText(
                f"| Оставшиеся часы: {current_error.hours} (= {current_error.hours // 2} {pair_text(current_error.hours // 2)})"
            )

    def event(self, event):  # noqa: ANN001
        if event.type() == QEvent.Type(124):
            self.show_help()
            return True
        return super().event(event)

    def show_help(self):
        # Отображение справки
        QMessageBox.information(
            self,
            QMessageBoxTitles.HELP.value,
            QMessageBoxMessages.ERROR_DIALOG_HELP.value,
        )


class MainWindow(QMainWindow):
    instance = None

    def __init__(self, data=None):  # noqa: ANN001
        MainWindow.instance = self
        self.data = data
        self.current_schedule: schedule_maker.Schedule = None
        self.remaining_data = None
        self.rating = None
        self.errors = []
        self.empty_table_message = "Здесь отобразится сгенерированное расписание"
        self.table_headers = [
            "Группа",
            "День",
            "Время",
            "Форма",
            "Предмет",
            "Педагог",
            "Каб.",
        ]
        self.current_cell = None
        self.input_dialog = None

        super().__init__()
        self._setup_window()
        self._setup_ui()
        self._load_stylesheet()
        self._initialize_data()

    def _setup_window(self) -> None:
        """Setup window properties."""
        self.setWindowTitle("Составление расписания")
        self.setWindowIcon(QIcon(db.resource_path("../icon.ico")))
        self.setGeometry(WINDOW_X, WINDOW_Y, 1400, 800)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QHBoxLayout(self.central_widget)

    def _setup_ui(self) -> None:
        """Setup UI components."""
        self._setup_table_widget()
        self._setup_buttons()
        self._setup_layout()

    def _setup_table_widget(self) -> None:
        """Setup the main table widget."""
        self.table_widget = QTableWidget(1, 1)
        self.table_widget.setHorizontalHeaderLabels(["Расписание"])
        self.table_widget.setItem(0, 0, QTableWidgetItem(self.empty_table_message))
        self.table_widget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_widget.cellClicked.connect(self.on_cell_click)

    def _setup_buttons(self) -> None:
        """Setup all buttons."""
        self.button_layout = QVBoxLayout()

        self.generate_schedule_button = QPushButton("Сгенерировать расписание")
        self.generate_schedule_button.clicked.connect(self.generate_schedule)
        self.button_layout.addWidget(self.generate_schedule_button)

        self.error_button = QPushButton(f"Ошибки: {len(self.errors)}")
        self.error_button.setEnabled(False)
        self.error_button.clicked.connect(self.show_errors)
        self.button_layout.addWidget(self.error_button)

        self.export_schedule_button = QPushButton("Выгрузить расписание")
        self.export_schedule_button.clicked.connect(self.export_schedule)
        self.button_layout.addWidget(self.export_schedule_button)

        self.input_data_button = QPushButton("Ввод данных")
        self.input_data_button.clicked.connect(self.input_data)
        self.button_layout.addWidget(self.input_data_button)

        self.sort_by_button = QPushButton("Сформировать по...")
        self.sort_by_button.clicked.connect(self.sort_by)
        self.sort_by_button.setEnabled(False)
        self.button_layout.addWidget(self.sort_by_button)

        self.generate_best_schedule_button = QPushButton(
            "Сгенерировать лучшее расписание"
        )
        self.generate_best_schedule_button.clicked.connect(self.generate_best_schedule)
        self.button_layout.addWidget(self.generate_best_schedule_button)

        self.schedule_rating_label = QLabel()
        self.button_layout.addWidget(self.schedule_rating_label)

    def _setup_layout(self) -> None:
        """Setup the main layout."""
        self.layout.addWidget(self.table_widget)
        self.layout.addLayout(self.button_layout)
        self.button_layout.addStretch()

    def _load_stylesheet(self) -> None:
        """Load the CSS stylesheet."""
        with Path(
            db.resource_path(f"../css/{self.__class__.__name__}.css")
        ).open() as f:
            self.setStyleSheet(f.read())
        self.resize_columns()

    def _initialize_data(self) -> None:
        """Initialize data if needed."""
        if self.data is None:
            resp = QMessageBox.warning(
                self,
                QMessageBoxTitles.INFORMATION.value,
                QMessageBoxMessages.NO_DATA_FOUND_TEMPLATE.value,
                QMessageBox.Ok | QMessageBox.Cancel,
            )

            if resp == QMessageBox.Ok:
                self.data = db.ExampleData()
            else:
                self.data = db.EmptyData()

    def schedule_rating_label_update(self, rating: dict[str, int]):
        names = {
            "rate": "Рейтинг",
            "teachers_gaps_count": "Окна у преподавателей",
            "offline_pairs_gaps": "Пропущенные пары",
            "overworked_teachers": "Перегруженные преподаватели",
            "unissued_hours": "Неиспользованные часы",
        }
        rating = {names[k]: v for k, v in rating.items()}
        self.schedule_rating_label.setText(
            f"Результат:\n - {'\n - '.join([f'{k}: {v}' for k, v in rating.items()])}"
        )

    def on_cell_click(self, row: int, column: int):
        sort_by = {
            "Группа": "group",
            "День": "day",
            "Форма": "pair_type",
            "Предмет": "discipline",
            "Педагог": "teacher",
            "Каб.": "classroom",
        }
        if self.current_cell == "all":
            return
        self.current_cell = {
            "row": row,
            "column": column,
            "value": self.table_widget.item(row, column).text(),
            "header": self.table_widget.horizontalHeaderItem(column).text(),
        }
        if self.current_cell["header"] in sort_by:
            self.sort_by_button.setText(
                f"Сформировать по\n{self.current_cell['value']}"
            )
            self.sort_by_button.setEnabled(True)
        else:
            self.sort_by_button.setText("Сформировать по...")
            self.sort_by_button.setEnabled(False)

    def resize_columns(self):
        self.table_widget.resizeColumnsToContents()

    def set_pairs_to_table(self, pairs: dict[str, list[db.Pair]]):
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(len(self.table_headers))
        self.table_widget.setHorizontalHeaderLabels(self.table_headers)

        rows: list[str] = []
        for group, pairs_list in pairs.items():
            rows.extend(
                [
                    group,
                    pair.day,
                    pair.pair_time.get_str(),
                    pair.pair_type,
                    pair.discipline,
                    pair.teacher,
                    pair.classroom,
                ]
                for pair in pairs_list
            )

        for row in rows:
            current_row = self.table_widget.rowCount()
            self.table_widget.insertRow(current_row)
            for column in range(len(row)):
                self.table_widget.setItem(
                    current_row, column, QTableWidgetItem(row[column])
                )

        self.resize_columns()

    def generate_schedule(self):
        start_time = time.time()
        logger.info("Starting schedule generation")
        working_data = copy.deepcopy(self.data)
        sch = schedule_maker.make_full_schedule(working_data)
        self.current_schedule = sch
        self.errors = sch.errors
        elapsed_time = time.time() - start_time
        logger.info(
            "Schedule generated, errors: %d, time: %.2fs",
            len(sch.errors),
            elapsed_time,
        )
        self.remaining_data = sch.remaining_data
        rating = {
            "rate": best_of.rate_schedule(sch.pairs, self.data, sch.remaining_data)
        } | best_of.get_counts(sch.pairs, self.data, sch.remaining_data)
        if self.rating is not None and self.rating["rate"] > rating["rate"]:
            resp = QMessageBox.question(
                self,
                QMessageBoxTitles.ATTENTION.value,
                QMessageBoxMessages.WORSE_SCHEDULE_QUESTION_TEMPLATE.value.format(
                    rating["rate"]
                ),
                buttons=QMessageBox.Ok | QMessageBox.Cancel,
            )
            if resp == QMessageBox.Cancel:
                return
        self.rating = rating
        self.schedule_rating_label_update(self.rating)
        self.set_pairs_to_table(sch.pairs)
        self.resize_columns()
        self.error_button.setText(f"Ошибки: {len(self.errors)}")
        self.error_button.setEnabled(len(self.errors) > 0)

    def show_errors(self):
        if self.errors:
            error_dialog = ErrorDialog(self.errors, self.remaining_data)
            error_dialog.exec_()

    def export_schedule(self):
        logger.info("Starting Excel export")

        try:
            file_name = self._generate_excel_file()
            self._show_export_dialog(file_name)
        except ExportError as e:
            logger.warning("Export error: %s", e)
            QMessageBox.warning(
                self,
                QMessageBoxTitles.ERROR.value,
                QMessageBoxMessages.SCHEDULE_NOT_GENERATED_ERROR.value,
            )
        except Exception as e:
            logger.exception("Error during schedule export")
            QMessageBox.warning(self, QMessageBoxTitles.ERROR.value, str(e))
            raise

    def _generate_excel_file(self) -> str:
        if self.current_schedule is None:
            msg = "Расписание не сформировано"
            raise ExportError(msg)

        row_count = self.table_widget.rowCount()
        column_count = self.table_widget.columnCount()

        file_name = "ExportSchedule{}{}.xlsx".format(
            int(self.rating["rate"]),
            datetime.datetime.now().strftime("%H%M%S"),
        )
        logger.info("Creating file: %s", file_name)

        exp_data = self._prepare_export_data(row_count, column_count)
        headers = self._get_export_headers(column_count)

        workbook = Workbook()
        sheet = workbook.active

        self._fill_excel_headers(sheet, headers)
        self._fill_excel_data(sheet, exp_data)
        self._fill_excel_errors(sheet, column_count)
        self._fill_excel_rating(sheet, column_count)
        self._adjust_column_width(sheet, column_count, row_count)

        workbook.save(file_name)
        return file_name

    def _prepare_export_data(
        self, row_count: int, column_count: int
    ) -> list[list[str]]:
        exp_data = []
        for row in range(row_count):
            row_data = []
            for column in range(column_count):
                item = self.table_widget.item(row, column)
                row_data.append(item.text() if item else "")
            exp_data.append(row_data)
        return exp_data

    def _get_export_headers(self, column_count: int) -> list[str]:
        return [
            self.table_widget.horizontalHeaderItem(i).text()
            for i in range(column_count)
        ]

    def _fill_excel_headers(self, sheet: Worksheet, headers: list[str]) -> None:
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num, value=header)

    def _fill_excel_data(self, sheet: Worksheet, exp_data: list[list[str]]) -> None:
        for row_num, row_data in enumerate(exp_data, 2):
            for col_num, value in enumerate(row_data, 1):
                sheet.cell(row=row_num, column=col_num, value=value)

    def _fill_excel_errors(self, sheet: Worksheet, column_count: int) -> None:
        sheet.cell(
            row=1,
            column=column_count + 2,
            value=f"Невозможно поставить пары: {len(self.errors)}",
        )
        err_headers = ["Группа", "Предмет", "Остаток часов"]
        for col_num, header in enumerate(err_headers, column_count + 2):
            sheet.cell(row=2, column=col_num, value=header)

        for err_num, err in enumerate(self.errors):
            sheet.cell(row=3 + err_num, column=column_count + 2, value=err.group)
            sheet.cell(row=3 + err_num, column=column_count + 3, value=err.discipline)
            sheet.cell(row=3 + err_num, column=column_count + 4, value=err.hours)

    def _fill_excel_rating(self, sheet: Worksheet, column_count: int) -> None:
        rating_info = [
            (len(self.errors) + 4, f"Рейтинг: {self.rating['rate']}"),
            (
                len(self.errors) + 5,
                f"Окна у преподавателей: {self.rating['teachers_gaps_count']}",
            ),
            (
                len(self.errors) + 6,
                f"Пропущенные пары: {self.rating['offline_pairs_gaps']}",
            ),
            (
                len(self.errors) + 7,
                f"Перегруженные преподаватели: {self.rating['overworked_teachers']}",
            ),
            (
                len(self.errors) + 8,
                f"Неиспользованные часы: {self.rating['unissued_hours']}",
            ),
        ]

        for row, text in rating_info:
            sheet.cell(row=row, column=column_count + 2, value=text)

    def _adjust_column_width(
        self, sheet: Worksheet, column_count: int, row_count: int
    ) -> None:
        err_headers = ["Группа", "Предмет", "Остаток часов"]
        for col in range(1, column_count + len(err_headers) + 4):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, row_count + 2):
                cell_value = sheet[f"{column_letter}{row}"].value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = max_length + EXCEL_COLUMN_WIDTH_PADDING
            sheet.column_dimensions[column_letter].width = adjusted_width

    def _show_export_dialog(self, file_name: str) -> None:
        resp = QMessageBox.information(
            self,
            QMessageBoxTitles.SUCCESS.value,
            QMessageBoxMessages.SCHEDULE_EXPORTED_SUCCESS.value,
            buttons=QMessageBox.Ok | QMessageBox.Cancel,
        )

        if resp == QMessageBox.Ok:
            logger.info("Opening file: %s", file_name)
            os.startfile(file_name)  # noqa: S606
        else:
            logger.info("File not opened by user choice")

    def sort_by(self):
        if self.current_cell == "all":
            self.set_pairs_to_table(self.current_schedule.pairs)
            self.current_cell = None
            self.sort_by_button.setText("Сформировать по...")
            self.sort_by_button.setEnabled(False)
            return
        if self.current_schedule is None:
            QMessageBox.warning(
                self,
                QMessageBoxTitles.ERROR.value,
                QMessageBoxMessages.GENERATE_BEFORE_SORT_ERROR.value,
            )
            return
        sort_by = {
            "Группа": "group",
            "День": "day",
            "Форма": "pair_type",
            "Предмет": "discipline",
            "Педагог": "teacher",
            "Каб.": "classroom",
        }
        sorted_pairs = schedule_maker.get_schedule_for(
            sort_by[self.current_cell["header"]],
            self.current_schedule.pairs,
            self.current_cell["value"],
        )
        sorted_pairs = schedule_maker.sorted_pairs(sorted_pairs)
        self.set_pairs_to_table(sorted_pairs)
        self.current_cell = "all"
        self.sort_by_button.setText("Полное расписание")
        self.sort_by_button.setEnabled(True)

    def generate_best_schedule(self):
        logger.info("Starting schedule optimization")
        dialog = ScheduleGeneratorDialog(self.data)
        dialog.result_obtained.connect(self.handle_generator_result)
        if dialog.exec_() == QDialog.Accepted:
            result, rating = dialog.get_result()
            self.handle_generator_result(result, rating)

    def input_data(self):
        if not self.input_dialog:  # Создаем окно только если оно еще не создано
            self.input_dialog = InputDataDialog(self.data)
        self.input_dialog.show()

    def handle_generator_result(
        self, result: schedule_maker.Schedule, rating: dict[str, int]
    ):
        logger.info("Generation result received, rating: %.2f", rating["rate"])
        if self.rating is not None and self.rating["rate"] > rating["rate"]:
            logger.warning("New schedule is worse than previous")
            resp = QMessageBox.question(
                self,
                QMessageBoxTitles.ATTENTION.value,
                QMessageBoxMessages.WORSE_SCHEDULE_QUESTION_TEMPLATE.value.format(
                    rating["rate"]
                ),
                buttons=QMessageBox.Ok | QMessageBox.Cancel,
            )
            if resp == QMessageBox.Cancel:
                return

        self.current_schedule = result
        self.errors = result.errors
        self.remaining_data = result.remaining_data
        self.rating = rating
        self.set_pairs_to_table(result.pairs)
        self.resize_columns()
        self.error_button.setText(f"Ошибки: {len(self.errors)}")
        self.error_button.setEnabled(len(self.errors) > 0)
        self.schedule_rating_label_update(self.rating)


def global_exception_handler(exctype: type, value: BaseException, _: traceback):
    logger.error("Unhandled exception occurred: %s", value)
    # Сохраняем в старый файл для совместимости
    with Path("./error_log.txt").open("a", encoding="utf-8") as f:
        f.write(
            "{} Unhandled exception occurred: {}\n\n".format(
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), value
            )
        )
    if (
        hasattr(MainWindow, "_instance")
        and MainWindow.instance is not None
        and hasattr(MainWindow.instance, "data")
        and MainWindow.instance.data
    ):
        try:
            db.save_data(MainWindow.instance.data)
            logger.info("Data automatically saved after error")
        except Exception as save_error:
            logger.exception("Failed to save data after error: %s", save_error)  # noqa: TRY401
    sys.__excepthook__(exctype, value, traceback)
    sys.exit(1)


sys.excepthook = global_exception_handler

if __name__ == "__main__":
    if db.check_exists_data():
        data = db.load_data()
        logger.info("Starting program ( #%d )", data.counter)
    else:
        data = db.ExampleData()
        logger.info("Starting program for the first time")
    app = QApplication(sys.argv)
    window = MainWindow(data)
    window.show()
    ex_code = app.exec_()
    db.save_data(data)
    sys.exit(ex_code)
